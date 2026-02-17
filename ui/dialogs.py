
"""
Dialogs for Flight Bot
"""
import sys
import logging
from datetime import datetime, timedelta
from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QComboBox, 
    QCalendarWidget, QGroupBox, QListWidget, QListWidgetItem, QFrame,
    QMessageBox, QDateEdit, QSpinBox, QCheckBox, QScrollArea, QGridLayout,
    QWidget, QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView,
    QTabWidget, QFileDialog, QInputDialog
)
from PyQt6.QtCore import Qt, pyqtSignal, QDate, QSettings
from PyQt6.QtGui import QColor, QFont, QTextCharFormat, QAction

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

import config
from ui.styles import MODERN_THEME
from ui.components import NoWheelSpinBox, NoWheelComboBox, NoWheelDateEdit

logger = logging.getLogger(__name__)

# --- Validation Helpers ---

def _validate_route_and_dates(parent, origin: str, dest: str, dep_date: QDate, ret_date: QDate = None) -> bool:
    """공통 노선/날짜 검증"""
    if origin == dest:
        QMessageBox.warning(parent, "입력 오류", "출발지와 도착지가 같습니다.")
        return False

    today = QDate.currentDate()
    if dep_date < today:
        QMessageBox.warning(parent, "날짜 오류", "출발일이 오늘보다 이전입니다.")
        return False

    if ret_date and ret_date < dep_date:
        QMessageBox.warning(parent, "날짜 오류", "귀국일이 출발일보다 이전입니다.")
        return False

    return True

# --- Calendar View Dialog ---

class CalendarViewDialog(QDialog):
    """월별 최저가 캘린더 뷰"""
    date_selected = pyqtSignal(str)  # 선택된 날짜 (yyyyMMdd)
    
    def __init__(self, price_data: dict, parent=None):
        """
        Args:
            price_data: {date_str: (min_price, airline)} 형식
        """
        super().__init__(parent)
        self.price_data = price_data
        self.setWindowTitle("📅 날짜별 최저가 캘린더")
        self.setMinimumSize(700, 550)
        self.setStyleSheet(MODERN_THEME)
        self._init_ui()
    
    def _init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        
        # 범례
        legend = QLabel("🟢 최저가  🟡 중간  🔴 비쌈  ⬜ 데이터 없음")
        legend.setStyleSheet("font-size: 12px; color: #94a3b8;")
        layout.addWidget(legend)
        
        # 캘린더 위젯
        self.calendar = QCalendarWidget()
        self.calendar.setGridVisible(True)
        self.calendar.setMinimumDate(QDate.currentDate())
        self.calendar.clicked.connect(self._on_date_clicked)
        layout.addWidget(self.calendar)
        
        # 가격 범위 계산
        if self.price_data:
            prices = [p for p, _ in self.price_data.values() if p > 0]
            if prices:
                self.min_price = min(prices)
                self.max_price = max(prices)
                self.price_range = self.max_price - self.min_price if self.max_price > self.min_price else 1
            else:
                self.min_price = self.max_price = self.price_range = 0
        else:
            self.min_price = self.max_price = self.price_range = 0
        
        # 날짜별 색상 적용
        self._apply_price_colors()
        
        # 선택된 날짜 정보
        self.lbl_info = QLabel("날짜를 클릭하면 해당 날짜로 검색합니다")
        self.lbl_info.setStyleSheet("font-size: 14px; font-weight: bold;")
        layout.addWidget(self.lbl_info)
        
        # 버튼
        btn_layout = QHBoxLayout()
        btn_close = QPushButton("닫기")
        btn_close.clicked.connect(self.close)
        btn_layout.addStretch()
        btn_layout.addWidget(btn_close)
        layout.addLayout(btn_layout)
    
    def _apply_price_colors(self):
        """날짜별 가격에 따른 색상 적용"""
        for date_str, (price, airline) in self.price_data.items():
            if price <= 0:
                continue
            
            # 날짜 파싱
            try:
                qdate = QDate.fromString(date_str, "yyyyMMdd")
                if not qdate.isValid():
                    continue
            except Exception as e:
                logger.debug(f"Date parsing error: {e}")
                continue
            
            # 가격 기반 색상 결정
            if self.price_range > 0:
                ratio = (price - self.min_price) / self.price_range
            else:
                ratio = 0
            
            if ratio < 0.3:
                color = QColor("#22c55e")  # 녹색 - 저렴
            elif ratio < 0.6:
                color = QColor("#f59e0b")  # 주황색 - 중간
            else:
                color = QColor("#ef4444")  # 빨간색 - 비쌈
            
            # 캘린더 날짜에 포맷 적용
            fmt = QTextCharFormat()
            fmt.setBackground(color)
            fmt.setForeground(QColor("white"))
            fmt.setToolTip(f"{price:,}원 ({airline})")
            self.calendar.setDateTextFormat(qdate, fmt)
    
    def _on_date_clicked(self, qdate):
        date_str = qdate.toString("yyyyMMdd")
        if date_str in self.price_data:
            price, airline = self.price_data[date_str]
            self.lbl_info.setText(f"📅 {qdate.toString('yyyy-MM-dd')}: {price:,}원 ({airline})")
        else:
            self.lbl_info.setText(f"📅 {qdate.toString('yyyy-MM-dd')}: 데이터 없음")
        
        # 시그널 발생
        self.date_selected.emit(date_str)


# --- Combination Selector Dialog ---

class CombinationSelectorDialog(QDialog):
    """가는편/오는편 개별 선택 다이얼로그"""
    combination_selected = pyqtSignal(object, object)  # outbound_flight, return_flight
    
    def __init__(self, outbound_flights: list, return_flights: list, parent=None):
        super().__init__(parent)
        self.outbound_flights = outbound_flights
        self.return_flights = return_flights
        self.selected_outbound = None
        self.selected_return = None
        
        self.setWindowTitle("✈️ 가는편/오는편 조합 선택")
        self.setMinimumSize(1000, 600)
        self.setStyleSheet(MODERN_THEME)
        self._init_ui()
    
    def _init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        
        # 상단 설명
        info = QLabel("가는편과 오는편을 각각 선택하여 원하는 조합을 만들 수 있습니다.")
        info.setStyleSheet("font-size: 13px; color: #94a3b8;")
        layout.addWidget(info)
        
        # 메인 컨텐츠 (좌: 가는편, 우: 오는편)
        content_layout = QHBoxLayout()
        
        # 가는편 리스트
        outbound_group = QGroupBox("✈️ 가는편 선택")
        outbound_layout = QVBoxLayout(outbound_group)
        
        self.list_outbound = QListWidget()
        self.list_outbound.setAlternatingRowColors(True)
        for i, flight in enumerate(self.outbound_flights):
            item_text = f"{flight.airline} | {flight.departure_time} → {flight.arrival_time} | {flight.price:,}원"
            item = QListWidgetItem(item_text)
            item.setData(Qt.ItemDataRole.UserRole, i)
            self.list_outbound.addItem(item)
        self.list_outbound.currentRowChanged.connect(self._on_outbound_selected)
        outbound_layout.addWidget(self.list_outbound)
        
        content_layout.addWidget(outbound_group)
        
        # 오는편 리스트
        return_group = QGroupBox("🔙 오는편 선택")
        return_layout = QVBoxLayout(return_group)
        
        self.list_return = QListWidget()
        self.list_return.setAlternatingRowColors(True)
        for i, flight in enumerate(self.return_flights):
            item_text = f"{flight.airline} | {flight.departure_time} → {flight.arrival_time} | {flight.price:,}원"
            item = QListWidgetItem(item_text)
            item.setData(Qt.ItemDataRole.UserRole, i)
            self.list_return.addItem(item)
        self.list_return.currentRowChanged.connect(self._on_return_selected)
        return_layout.addWidget(self.list_return)
        
        content_layout.addWidget(return_group)
        
        layout.addLayout(content_layout)
        
        # 선택된 조합 요약
        summary_frame = QFrame()
        summary_frame.setStyleSheet("background-color: #16213e; border-radius: 8px; padding: 15px;")
        summary_layout = QVBoxLayout(summary_frame)
        
        self.lbl_summary = QLabel("가는편과 오는편을 선택하세요")
        self.lbl_summary.setStyleSheet("font-size: 16px; font-weight: bold;")
        summary_layout.addWidget(self.lbl_summary)
        
        self.lbl_total = QLabel("")
        self.lbl_total.setStyleSheet("font-size: 24px; font-weight: bold; color: #22c55e;")
        summary_layout.addWidget(self.lbl_total)
        
        layout.addWidget(summary_frame)
        
        # 버튼
        btn_layout = QHBoxLayout()
        
        btn_confirm = QPushButton("✅ 이 조합으로 선택")
        btn_confirm.setEnabled(False)
        btn_confirm.clicked.connect(self._on_confirm)
        self.btn_confirm = btn_confirm
        
        btn_cancel = QPushButton("취소")
        btn_cancel.clicked.connect(self.close)
        
        btn_layout.addStretch()
        btn_layout.addWidget(btn_confirm)
        btn_layout.addWidget(btn_cancel)
        
        layout.addLayout(btn_layout)
    
    def _on_outbound_selected(self, row):
        if 0 <= row < len(self.outbound_flights):
            self.selected_outbound = self.outbound_flights[row]
            self._update_summary()
    
    def _on_return_selected(self, row):
        if 0 <= row < len(self.return_flights):
            self.selected_return = self.return_flights[row]
            self._update_summary()
    
    def _update_summary(self):
        if self.selected_outbound and self.selected_return:
            out = self.selected_outbound
            ret = self.selected_return
            
            total = out.price + ret.price
            
            self.lbl_summary.setText(
                f"가는편: {out.airline} {out.departure_time}→{out.arrival_time} ({out.price:,}원)\\n"
                f"오는편: {ret.airline} {ret.departure_time}→{ret.arrival_time} ({ret.price:,}원)"
            )
            self.lbl_total.setText(f"총 {total:,}원")
            self.btn_confirm.setEnabled(True)
        elif self.selected_outbound:
            self.lbl_summary.setText(f"가는편 선택됨: {self.selected_outbound.airline} - 오는편을 선택하세요")
            self.lbl_total.setText("")
            self.btn_confirm.setEnabled(False)
        elif self.selected_return:
            self.lbl_summary.setText(f"오는편 선택됨: {self.selected_return.airline} - 가는편을 선택하세요")
            self.lbl_total.setText("")
            self.btn_confirm.setEnabled(False)
    
    def _on_confirm(self):
        if self.selected_outbound and self.selected_return:
            self.combination_selected.emit(self.selected_outbound, self.selected_return)
            self.accept()

# --- Dialogs ---

class MultiDestDialog(QDialog):
    """다중 목적지 선택 다이얼로그"""
    search_requested = pyqtSignal(str, list, str, str, int)  # origin, dests, dep, ret, adults
    
    def __init__(self, parent=None, prefs=None):
        super().__init__(parent)
        self.prefs = prefs
        self.setWindowTitle("🌍 다중 목적지 검색")
        self.setMinimumSize(500, 500)
        self.setStyleSheet(MODERN_THEME)
        self._init_ui()
    
    def _init_ui(self):
        layout = QVBoxLayout(self)
        
        # Instructions
        layout.addWidget(QLabel("여러 목적지를 선택하여 한 번에 비교 검색합니다."))
        
        # Origin
        origin_layout = QHBoxLayout()
        origin_layout.addWidget(QLabel("출발지:"))
        self.cb_origin = QComboBox()
        for code, name in config.AIRPORTS.items():
            self.cb_origin.addItem(f"{code} ({name})", code)
        self.cb_origin.setCurrentIndex(0)
        origin_layout.addWidget(self.cb_origin)
        layout.addLayout(origin_layout)
        
        # Destination Checkboxes
        layout.addWidget(QLabel("도착지 선택 (다중 선택 가능):"))
        
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        dest_widget = QWidget()
        dest_layout = QGridLayout(dest_widget)
        
        self.dest_checkboxes = {}
        all_presets = self.prefs.get_all_presets() if self.prefs else config.AIRPORTS
        
        row, col = 0, 0
        for code, name in all_presets.items():
            cb = QCheckBox(f"{code} ({name})")
            cb.setProperty("code", code)
            self.dest_checkboxes[code] = cb
            dest_layout.addWidget(cb, row, col)
            col += 1
            if col >= 3:
                col = 0
                row += 1
        
        scroll.setWidget(dest_widget)
        layout.addWidget(scroll, 1)
        
        # Select All / None
        btn_layout = QHBoxLayout()
        btn_all = QPushButton("모두 선택")
        btn_all.clicked.connect(lambda: self._toggle_all(True))
        btn_none = QPushButton("모두 해제")
        btn_none.clicked.connect(lambda: self._toggle_all(False))
        btn_layout.addWidget(btn_all)
        btn_layout.addWidget(btn_none)
        layout.addLayout(btn_layout)
        
        # Dates
        date_layout = QHBoxLayout()
        date_layout.addWidget(QLabel("가는 날:"))
        self.date_dep = QDateEdit()
        self.date_dep.setCalendarPopup(True)
        self.date_dep.setDate(QDate.currentDate().addDays(7))
        date_layout.addWidget(self.date_dep)
        
        date_layout.addWidget(QLabel("오는 날:"))
        self.date_ret = QDateEdit()
        self.date_ret.setCalendarPopup(True)
        self.date_ret.setDate(QDate.currentDate().addDays(10))
        date_layout.addWidget(self.date_ret)
        layout.addLayout(date_layout)
        
        # Adults
        adult_layout = QHBoxLayout()
        adult_layout.addWidget(QLabel("성인:"))
        self.spin_adults = QSpinBox()
        self.spin_adults.setRange(1, 9)
        adult_layout.addWidget(self.spin_adults)
        adult_layout.addStretch()
        layout.addLayout(adult_layout)
        
        # Action Buttons (Cancel left, Action right - UX standard)
        action_layout = QHBoxLayout()
        action_layout.addStretch()
        
        btn_cancel = QPushButton("❌ 취소")
        btn_cancel.setObjectName("secondary_btn")
        btn_cancel.clicked.connect(self.reject)
        action_layout.addWidget(btn_cancel)
        
        btn_search = QPushButton("🔍 다중 검색 시작")
        btn_search.clicked.connect(self._on_search)
        action_layout.addWidget(btn_search)
        
        layout.addLayout(action_layout)
    
    def _toggle_all(self, checked):
        for cb in self.dest_checkboxes.values():
            cb.setChecked(checked)
    
    def _on_search(self):
        selected = [code for code, cb in self.dest_checkboxes.items() if cb.isChecked()]
        
        if len(selected) < 2:
            QMessageBox.warning(self, "선택 오류", "최소 2개 이상의 목적지를 선택하세요.")
            return
        
        origin = self.cb_origin.currentData()
        dep_date = self.date_dep.date()
        ret_date = self.date_ret.date()

        if origin in selected:
            QMessageBox.warning(self, "선택 오류", "출발지는 도착지 목록에 포함할 수 없습니다.")
            return

        if not _validate_route_and_dates(self, origin, selected[0], dep_date, ret_date):
            return

        dep = dep_date.toString("yyyyMMdd")
        ret = ret_date.toString("yyyyMMdd")
        adults = self.spin_adults.value()
        
        self.search_requested.emit(origin, selected, dep, ret, adults)
        self.accept()


class DateRangeDialog(QDialog):
    """날짜 범위 검색 다이얼로그"""
    search_requested = pyqtSignal(str, str, list, int, int)  # origin, dest, dates, return_offset, adults
    
    def __init__(self, parent=None, prefs=None):
        super().__init__(parent)
        self.prefs = prefs
        self.setWindowTitle("📅 날짜 범위 검색")
        self.setMinimumSize(450, 400)
        self.setStyleSheet(MODERN_THEME)
        self._init_ui()
    
    def _init_ui(self):
        layout = QVBoxLayout(self)
        
        layout.addWidget(QLabel("날짜 범위를 지정하여 가장 저렴한 날짜를 찾습니다."))
        
        # Origin & Dest
        route_layout = QHBoxLayout()
        route_layout.addWidget(QLabel("출발지:"))
        self.cb_origin = QComboBox()
        for code, name in config.AIRPORTS.items():
            self.cb_origin.addItem(f"{code} ({name})", code)
        route_layout.addWidget(self.cb_origin)
        
        route_layout.addWidget(QLabel("→"))
        
        route_layout.addWidget(QLabel("도착지:"))
        self.cb_dest = QComboBox()
        all_presets = self.prefs.get_all_presets() if self.prefs else config.AIRPORTS
        for code, name in all_presets.items():
            self.cb_dest.addItem(f"{code} ({name})", code)
        self.cb_dest.setCurrentIndex(1)  # 두 번째 항목
        route_layout.addWidget(self.cb_dest)
        layout.addLayout(route_layout)
        
        # Date Range
        layout.addWidget(QLabel("검색 날짜 범위:"))
        date_layout = QHBoxLayout()
        
        date_layout.addWidget(QLabel("시작:"))
        self.date_start = QDateEdit()
        self.date_start.setCalendarPopup(True)
        self.date_start.setDate(QDate.currentDate().addDays(7))
        date_layout.addWidget(self.date_start)
        
        date_layout.addWidget(QLabel("종료:"))
        self.date_end = QDateEdit()
        self.date_end.setCalendarPopup(True)
        self.date_end.setDate(QDate.currentDate().addDays(14))
        date_layout.addWidget(self.date_end)
        layout.addLayout(date_layout)
        
        # Trip Duration
        dur_layout = QHBoxLayout()
        dur_layout.addWidget(QLabel("여행 기간:"))
        self.spin_duration = QSpinBox()
        self.spin_duration.setRange(0, 30)
        self.spin_duration.setValue(3)
        self.spin_duration.setSuffix("박")
        dur_layout.addWidget(self.spin_duration)
        dur_layout.addWidget(QLabel("(0 = 편도)"))
        dur_layout.addStretch()
        layout.addLayout(dur_layout)
        
        # Adults
        adult_layout = QHBoxLayout()
        adult_layout.addWidget(QLabel("성인:"))
        self.spin_adults = QSpinBox()
        self.spin_adults.setRange(1, 9)
        adult_layout.addWidget(self.spin_adults)
        adult_layout.addStretch()
        layout.addLayout(adult_layout)
        
        # Note
        note = QLabel("⚠️ 날짜 범위가 넓을수록 검색 시간이 오래 걸립니다.")
        note.setStyleSheet("color: #f0ad4e; font-size: 12px;")
        layout.addWidget(note)
        
        layout.addStretch()
        
        # Actions
        action_layout = QHBoxLayout()
        btn_search = QPushButton("🔍 날짜 검색 시작")
        btn_search.clicked.connect(self._on_search)
        btn_cancel = QPushButton("취소")
        btn_cancel.clicked.connect(self.reject)
        action_layout.addWidget(btn_search)
        action_layout.addWidget(btn_cancel)
        layout.addLayout(action_layout)
    
    def _on_search(self):
        start = self.date_start.date()
        end = self.date_end.date()
        today = QDate.currentDate()
        origin = self.cb_origin.currentData()
        dest = self.cb_dest.currentData()
        duration = self.spin_duration.value()
        adults = self.spin_adults.value()
        
        if origin == dest:
            QMessageBox.warning(self, "입력 오류", "출발지와 도착지가 같습니다.")
            return

        if start < today:
            QMessageBox.warning(self, "날짜 오류", "시작 날짜가 오늘보다 이전입니다.")
            return

        if end < start:
            QMessageBox.warning(self, "날짜 오류", "종료 날짜는 시작 날짜와 같거나 이후여야 합니다.")
            return

        if duration > 0:
            max_return = end.addDays(duration)
            if max_return < today:
                QMessageBox.warning(self, "날짜 오류", "여행 기간을 포함한 귀국일이 오늘보다 이전입니다.")
                return

        # Generate date list (start == end 허용)
        dates = []
        current = start
        while current <= end:
            dates.append(current.toString("yyyyMMdd"))
            current = current.addDays(1)

        if len(dates) > 14:
            reply = QMessageBox.question(
                self, "확인",
                f"{len(dates)}일을 검색합니다. 시간이 오래 걸릴 수 있습니다.\\n계속하시겠습니까?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply != QMessageBox.StandardButton.Yes:
                return
        
        self.search_requested.emit(origin, dest, dates, duration, adults)
        self.accept()


class MultiDestResultDialog(QDialog):
    """다중 목적지 검색 결과 비교 다이얼로그"""
    
    def __init__(self, results: dict, parent=None):
        super().__init__(parent)
        self.results = results  # {dest: [FlightResult]}
        self.setWindowTitle("🌍 다중 목적지 비교 결과")
        self.setMinimumSize(700, 500)
        self.setStyleSheet(MODERN_THEME)
        self._init_ui()
    
    def _init_ui(self):
        layout = QVBoxLayout(self)
        
        # Summary Table
        layout.addWidget(QLabel("목적지별 최저가 비교:", objectName="section_title"))
        
        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(["목적지", "최저가", "항공사", "출발시간", "결과 수"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.setAlternatingRowColors(True)
        
        # Sort by lowest price
        sorted_results = sorted(
            self.results.items(), 
            key=lambda x: min(r.price for r in x[1]) if x[1] else float('inf')
        )
        
        self.table.setRowCount(len(sorted_results))
        
        for i, (dest, flights) in enumerate(sorted_results):
            dest_name = config.AIRPORTS.get(dest, dest)
            self.table.setItem(i, 0, QTableWidgetItem(f"{dest} ({dest_name})"))
            
            if flights:
                best = min(flights, key=lambda x: x.price)
                
                price_item = QTableWidgetItem(f"{best.price:,}원")
                price_item.setForeground(QColor("#4cc9f0"))
                self.table.setItem(i, 1, price_item)
                self.table.setItem(i, 2, QTableWidgetItem(best.airline))
                self.table.setItem(i, 3, QTableWidgetItem(best.departure_time))
                self.table.setItem(i, 4, QTableWidgetItem(str(len(flights))))
            else:
                self.table.setItem(i, 1, QTableWidgetItem("N/A"))
                self.table.setItem(i, 2, QTableWidgetItem("-"))
                self.table.setItem(i, 3, QTableWidgetItem("-"))
                self.table.setItem(i, 4, QTableWidgetItem("0"))
        
        layout.addWidget(self.table)
        
        # Best recommendation
        if sorted_results and sorted_results[0][1]:
            best_dest = sorted_results[0][0]
            best_price = min(r.price for r in sorted_results[0][1])
            rec_label = QLabel(f"💡 추천: {best_dest} ({config.AIRPORTS.get(best_dest, '')}) - {best_price:,}원")
            rec_label.setStyleSheet("font-size: 16px; color: #4cc9f0; font-weight: bold; padding: 10px;")
            layout.addWidget(rec_label)
        
        # Close button
        btn_close = QPushButton("닫기")
        btn_close.clicked.connect(self.accept)
        layout.addWidget(btn_close)


class DateRangeResultDialog(QDialog):
    """날짜 범위 검색 결과 다이얼로그"""
    
    def __init__(self, results: dict, parent=None):
        super().__init__(parent)
        self.results = results  # {date: (price, airline)}
        self.setWindowTitle("📅 날짜별 최저가 결과")
        self.setMinimumSize(600, 500)
        self.setStyleSheet(MODERN_THEME)
        self._init_ui()
    
    def _init_ui(self):
        layout = QVBoxLayout(self)
        
        layout.addWidget(QLabel("날짜별 최저가:", objectName="section_title"))
        
        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["날짜", "요일", "최저가", "항공사"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.setAlternatingRowColors(True)
        
        sorted_dates = sorted(self.results.items(), key=lambda x: x[0])
        self.table.setRowCount(len(sorted_dates))
        
        min_price = min((p for p, a in self.results.values() if p > 0), default=0)
        
        weekdays = ["월", "화", "수", "목", "금", "토", "일"]
        
        for i, (date, (price, airline)) in enumerate(sorted_dates):
            # Format date
            try:
                dt = datetime.strptime(date, "%Y%m%d")
                date_str = dt.strftime("%Y-%m-%d")
                weekday = weekdays[dt.weekday()]
            except:
                date_str = date
                weekday = "-"
            
            self.table.setItem(i, 0, QTableWidgetItem(date_str))
            self.table.setItem(i, 1, QTableWidgetItem(weekday))
            
            if price > 0:
                price_item = QTableWidgetItem(f"{price:,}원")
                if price == min_price:
                    price_item.setForeground(QColor("#00ff00"))
                    price_item.setFont(QFont("Pretendard", 10, QFont.Weight.Bold))
                else:
                    price_item.setForeground(QColor("#4cc9f0"))
                self.table.setItem(i, 2, price_item)
                self.table.setItem(i, 3, QTableWidgetItem(airline))
            else:
                self.table.setItem(i, 2, QTableWidgetItem("N/A"))
                self.table.setItem(i, 3, QTableWidgetItem("-"))
        
        layout.addWidget(self.table)
        
        # Best date
        valid_results = [(d, p, a) for d, (p, a) in self.results.items() if p > 0]
        if valid_results:
            best = min(valid_results, key=lambda x: x[1])
            try:
                best_dt = datetime.strptime(best[0], "%Y%m%d")
                best_str = best_dt.strftime("%Y-%m-%d (%a)")
            except Exception as e:
                logger.debug(f"Date format error: {e}")
                best_str = best[0]
            rec_label = QLabel(f"💡 최저가 날짜: {best_str} - {best[1]:,}원 ({best[2]})")
            rec_label.setStyleSheet("font-size: 16px; color: #00ff00; font-weight: bold; padding: 10px;")
            layout.addWidget(rec_label)
        
        btn_close = QPushButton("닫기")
        btn_close.clicked.connect(self.accept)
        layout.addWidget(btn_close)


class ShortcutsDialog(QDialog):
    """키보드 단축키 도움말 다이얼로그"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("⌨️ 키보드 단축키")
        self.setMinimumSize(400, 300)
        self.setStyleSheet(MODERN_THEME)
        self._init_ui()
    
    def _init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        
        title = QLabel("키보드 단축키 목록")
        title.setStyleSheet("font-size: 18px; font-weight: bold; color: #4cc9f0;")
        layout.addWidget(title)
        
        shortcuts = [
            ("Ctrl + Enter", "검색 시작"),
            ("F5", "결과 새로고침 (필터 재적용)"),
            ("Escape", "검색 취소 / 다이얼로그 닫기"),
            ("Ctrl + F", "필터 항공사 선택으로 이동"),
            ("더블클릭", "결과 행에서 예약 페이지 열기"),
            ("우클릭", "결과 행에서 컨텍스트 메뉴"),
        ]
        
        for key, desc in shortcuts:
            row = QHBoxLayout()
            key_label = QLabel(key)
            key_label.setStyleSheet("""
                background-color: #0f3460;
                color: #4cc9f0;
                padding: 5px 10px;
                border-radius: 4px;
                font-family: 'Consolas', monospace;
                font-weight: bold;
            """)
            key_label.setFixedWidth(120)
            row.addWidget(key_label)
            
            desc_label = QLabel(desc)
            desc_label.setStyleSheet("color: #e6e6e6;")
            row.addWidget(desc_label)
            row.addStretch()
            
            layout.addLayout(row)
        
        layout.addStretch()
        
        btn_close = QPushButton("닫기")
        btn_close.clicked.connect(self.accept)
        layout.addWidget(btn_close)


class PriceAlertDialog(QDialog):
    """가격 알림 설정 다이얼로그"""
    
    def __init__(self, parent=None, db=None, prefs=None):
        super().__init__(parent)
        self.db = db
        self.prefs = prefs
        self.setWindowTitle("🔔 가격 알림 관리")
        self.setMinimumSize(700, 550)
        self.setStyleSheet(MODERN_THEME)
        self._init_ui()
        self._refresh_alerts()
    
    def _init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        
        # 설명
        info = QLabel("목표 가격을 설정하면 가격이 그 이하로 떨어질 때 알림을 받을 수 있습니다.")
        info.setStyleSheet("color: #94a3b8; font-size: 12px;")
        layout.addWidget(info)
        
        # 새 알림 추가 그룹
        grp_new = QGroupBox("➕ 새 알림 추가")
        new_layout = QGridLayout(grp_new)
        
        # 출발지
        new_layout.addWidget(QLabel("출발지:"), 0, 0)
        self.cb_origin = QComboBox()
        for code, name in config.AIRPORTS.items():
            self.cb_origin.addItem(f"{code} ({name})", code)
        new_layout.addWidget(self.cb_origin, 0, 1)
        
        # 도착지
        new_layout.addWidget(QLabel("도착지:"), 0, 2)
        self.cb_dest = QComboBox()
        all_presets = self.prefs.get_all_presets() if self.prefs else config.AIRPORTS
        for code, name in all_presets.items():
            self.cb_dest.addItem(f"{code} ({name})", code)
        self.cb_dest.setCurrentIndex(1)
        new_layout.addWidget(self.cb_dest, 0, 3)
        
        # 가는 날
        new_layout.addWidget(QLabel("가는 날:"), 1, 0)
        self.date_dep = QDateEdit()
        self.date_dep.setCalendarPopup(True)
        self.date_dep.setDate(QDate.currentDate().addDays(7))
        new_layout.addWidget(self.date_dep, 1, 1)
        
        # 오는 날
        new_layout.addWidget(QLabel("오는 날:"), 1, 2)
        self.date_ret = QDateEdit()
        self.date_ret.setCalendarPopup(True)
        self.date_ret.setDate(QDate.currentDate().addDays(10))
        new_layout.addWidget(self.date_ret, 1, 3)

        # 편도 알림
        self.chk_oneway = QCheckBox("편도 알림")
        self.chk_oneway.setToolTip("체크하면 귀국일 없이 편도 노선만 기준으로 알림을 설정합니다.")
        self.chk_oneway.toggled.connect(self._toggle_alert_oneway)
        new_layout.addWidget(self.chk_oneway, 2, 0, 1, 2)
        
        # 목표 가격
        new_layout.addWidget(QLabel("목표 가격:"), 3, 0)
        self.spin_target = QSpinBox()
        self.spin_target.setRange(10000, 10000000)
        self.spin_target.setSingleStep(10000)
        self.spin_target.setValue(300000)
        self.spin_target.setSuffix(" 원")
        new_layout.addWidget(self.spin_target, 3, 1)
        
        # 추가 버튼
        btn_add = QPushButton("🔔 알림 추가")
        btn_add.clicked.connect(self._add_alert)
        new_layout.addWidget(btn_add, 3, 2, 1, 2)
        
        layout.addWidget(grp_new)
        
        # 현재 알림 목록
        layout.addWidget(QLabel("📋 설정된 알림 목록:", objectName="section_title"))
        
        self.table = QTableWidget()
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels([
            "ID", "노선", "출발일", "귀국일", "목표가", "현재가", "상태"
        ])
        self.table.setColumnHidden(0, True)  # ID 숨김
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        layout.addWidget(self.table)
        
        # 버튼 영역
        btn_layout = QHBoxLayout()
        
        btn_refresh = QPushButton("🔄 새로고침")
        btn_refresh.clicked.connect(self._refresh_alerts)
        btn_layout.addWidget(btn_refresh)
        
        btn_delete = QPushButton("🗑️ 선택 삭제")
        btn_delete.clicked.connect(self._delete_selected)
        btn_layout.addWidget(btn_delete)
        
        btn_layout.addStretch()
        
        btn_close = QPushButton("닫기")
        btn_close.clicked.connect(self.accept)
        btn_layout.addWidget(btn_close)
        
        layout.addLayout(btn_layout)

    def _toggle_alert_oneway(self, checked: bool):
        self.date_ret.setEnabled(not checked)
    
    def _add_alert(self):
        """새 알림 추가"""
        origin = self.cb_origin.currentData()
        dest = self.cb_dest.currentData()
        dep_date = self.date_dep.date()
        ret_date = None if self.chk_oneway.isChecked() else self.date_ret.date()
        target = self.spin_target.value()

        if not _validate_route_and_dates(self, origin, dest, dep_date, ret_date):
            return

        dep = dep_date.toString("yyyyMMdd")
        ret = ret_date.toString("yyyyMMdd") if ret_date else None

        try:
            alert_id = self.db.add_price_alert(origin, dest, dep, ret, target)
            QMessageBox.information(self, "완료", f"가격 알림이 추가되었습니다. (ID: {alert_id})")
            self._refresh_alerts()
        except Exception as e:
            QMessageBox.critical(self, "오류", f"알림 추가 실패: {e}")
    
    def _refresh_alerts(self):
        """알림 목록 새로고침"""
        if not self.db:
            return
        
        alerts = self.db.get_all_alerts()
        self.table.setRowCount(len(alerts))
        
        for i, alert in enumerate(alerts):
            self.table.setItem(i, 0, QTableWidgetItem(str(alert.id)))
            
            route = f"{alert.origin} → {alert.destination}"
            self.table.setItem(i, 1, QTableWidgetItem(route))
            
            # 날짜 포맷
            try:
                dep_dt = datetime.strptime(alert.departure_date, "%Y%m%d")
                dep_str = dep_dt.strftime("%Y-%m-%d")
            except:
                dep_str = alert.departure_date
            self.table.setItem(i, 2, QTableWidgetItem(dep_str))
            
            if alert.return_date:
                try:
                    ret_dt = datetime.strptime(alert.return_date, "%Y%m%d")
                    ret_str = ret_dt.strftime("%Y-%m-%d")
                except:
                    ret_str = alert.return_date
            else:
                ret_str = "-"
            self.table.setItem(i, 3, QTableWidgetItem(ret_str))
            
            # 목표 가격
            target_item = QTableWidgetItem(f"{alert.target_price:,}원")
            target_item.setForeground(QColor("#4cc9f0"))
            self.table.setItem(i, 4, target_item)
            
            # 현재 가격
            if alert.last_price:
                current_item = QTableWidgetItem(f"{alert.last_price:,}원")
                if alert.last_price <= alert.target_price:
                    current_item.setForeground(QColor("#22c55e"))
                else:
                    current_item.setForeground(QColor("#f59e0b"))
            else:
                current_item = QTableWidgetItem("미확인")
            self.table.setItem(i, 5, current_item)
            
            # 상태
            if alert.triggered:
                status = "✅ 발동됨"
                color = "#22c55e"
            elif alert.is_active:
                status = "🔔 활성"
                color = "#4cc9f0"
            else:
                status = "⏸️ 비활성"
                color = "#94a3b8"
            
            status_item = QTableWidgetItem(status)
            status_item.setForeground(QColor(color))
            self.table.setItem(i, 6, status_item)
    
    def _delete_selected(self):
        """선택된 알림 삭제"""
        row = self.table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "선택 오류", "삭제할 알림을 선택하세요.")
            return
        
        alert_id = int(self.table.item(row, 0).text())
        reply = QMessageBox.question(
            self, "삭제 확인",
            "선택한 알림을 삭제하시겠습니까?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            if self.db.delete_alert(alert_id):
                self._refresh_alerts()
            else:
                QMessageBox.critical(self, "오류", "알림 삭제에 실패했습니다.")


class SettingsDialog(QDialog):
    def __init__(self, parent=None, prefs=None):
        super().__init__(parent)
        self.prefs = prefs
        self.setWindowTitle("⚙️ 설정 (Settings)")
        self.setMinimumSize(600, 500)  # Increased size for better content display
        self.setStyleSheet(MODERN_THEME)
        self._init_ui()
        
    def _init_ui(self):
        layout = QVBoxLayout(self)
        
        tabs = QTabWidget()
        tabs.addTab(self._create_general_tab(), "일반")
        tabs.addTab(self._create_dest_tab(), "목적지 관리")
        tabs.addTab(self._create_data_tab(), "데이터 관리")
        
        layout.addWidget(tabs)
        
        btn_close = QPushButton("닫기")
        btn_close.clicked.connect(self.accept)
        layout.addWidget(btn_close)
        
    def _create_general_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Preferred Time
        grp_time = QGroupBox("기본 선호 시간대")
        gt_layout = QHBoxLayout(grp_time)
        
        self.spin_start = QSpinBox()
        self.spin_start.setRange(0, 23)
        self.spin_end = QSpinBox()
        self.spin_end.setRange(1, 24)
        
        pt = self.prefs.get_preferred_time()
        self.spin_start.setValue(pt.get("departure_start", 0))
        self.spin_end.setValue(pt.get("departure_end", 24))
        
        gt_layout.addWidget(QLabel("출발:"))
        gt_layout.addWidget(self.spin_start)
        gt_layout.addWidget(QLabel("~"))
        gt_layout.addWidget(self.spin_end)
        gt_layout.addWidget(QLabel("시"))
        
        # Max Results
        grp_limit = QGroupBox("검색 결과 제한")
        gl_layout = QHBoxLayout(grp_limit)
        
        self.spin_limit = QSpinBox()
        self.spin_limit.setRange(50, 2000)
        self.spin_limit.setSingleStep(50)
        self.spin_limit.setValue(self.prefs.get_max_results())
        self.spin_limit.setSuffix(" 개")
        self.spin_limit.setToolTip("한 번의 검색에서 표시할 최대 결과 수 (기본: 1000)")
        
        gl_layout.addWidget(QLabel("최대 표시 개수:"))
        gl_layout.addWidget(self.spin_limit)
        gl_layout.addStretch()
        
        # Save Button (Combined)
        btn_save_time = QPushButton("설정 저장")
        btn_save_time.setFixedWidth(100)  # Increased width for visibility
        btn_save_time.setFixedHeight(30)
        btn_save_time.clicked.connect(self._save_time_pref)
        
        gt_layout.addWidget(btn_save_time)
        
        layout.addWidget(grp_time)
        layout.addWidget(grp_limit)
        layout.addStretch()
        return widget

    def _save_time_pref(self):
        self.prefs.set_preferred_time(self.spin_start.value(), self.spin_end.value())
        self.prefs.set_max_results(self.spin_limit.value())
        QMessageBox.information(self, "저장", "설정이 저장되었습니다.")

    def _create_dest_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        layout.addWidget(QLabel("사용자 정의 프리셋 목록:"))
        self.list_presets = QListWidget()
        self._refresh_presets()
        layout.addWidget(self.list_presets)
        
        btn_del = QPushButton("선택 삭제")
        btn_del.clicked.connect(self._delete_preset)
        layout.addWidget(btn_del)
        
        return widget
        
    def _refresh_presets(self):
        self.list_presets.clear()
        presets = self.prefs.get_all_presets()
        for code, name in presets.items():
            if code not in config.AIRPORTS: # Only show custom ones
                self.list_presets.addItem(f"{code} - {name}")

    def _delete_preset(self):
        item = self.list_presets.currentItem()
        if item:
            code = item.text().split(' - ')[0]
            self.prefs.remove_preset(code)
            self._refresh_presets()

    def _create_data_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # 프로필 가져오기/내보내기 그룹
        grp_profile = QGroupBox("📦 설정 백업/복원")
        gp_layout = QVBoxLayout(grp_profile)
        
        btn_export_settings = QPushButton("💾 모든 설정 내보내기")
        btn_export_settings.setToolTip("프리셋, 프로필, 선호 시간대, 테마 등 모든 설정을 JSON 파일로 저장")
        btn_export_settings.clicked.connect(self._export_all_settings)
        
        btn_import_settings = QPushButton("📂 설정 가져오기")
        btn_import_settings.setToolTip("JSON 파일에서 설정을 불러와 현재 설정에 병합")
        btn_import_settings.clicked.connect(self._import_all_settings)
        
        gp_layout.addWidget(btn_export_settings)
        gp_layout.addWidget(btn_import_settings)
        layout.addWidget(grp_profile)
        
        # 엑셀 그룹
        grp_excel = QGroupBox("📊 엑셀 (Excel)")
        gl = QVBoxLayout(grp_excel)
        
        btn_import = QPushButton("📂 검색 조건 가져오기 (Import)")
        btn_import.clicked.connect(self._import_excel)
        
        label_info = QLabel("엑셀 파일 양식: Origin, Dest, DepDate(YYYYMMDD), RetDate, Adults")
        label_info.setStyleSheet("font-size: 11px; color: #aaa;")
        
        btn_export = QPushButton("💾 검색 결과 내보내기 (Export)")
        btn_export.clicked.connect(self._export_excel)
        
        gl.addWidget(btn_import)
        gl.addWidget(btn_export)
        gl.addWidget(label_info)
        
        layout.addWidget(grp_excel)
        layout.addStretch()
        return widget

    def _export_all_settings(self):
        """모든 설정을 JSON 파일로 내보내기"""
        from datetime import datetime
        filename, _ = QFileDialog.getSaveFileName(
            self, "설정 내보내기",
            f"flight_bot_settings_{datetime.now().strftime('%Y%m%d')}.json",
            "JSON Files (*.json)"
        )
        if not filename:
            return
        
        if self.prefs.export_all_settings(filename):
            QMessageBox.information(self, "완료", f"설정이 저장되었습니다:\\n{filename}")
        else:
            QMessageBox.critical(self, "오류", "설정 내보내기에 실패했습니다.")

    def _import_all_settings(self):
        """JSON 파일에서 설정 가져오기"""
        filename, _ = QFileDialog.getOpenFileName(
            self, "설정 가져오기",
            "",
            "JSON Files (*.json)"
        )
        if not filename:
            return
        
        reply = QMessageBox.question(
            self, "설정 가져오기",
            "현재 설정에 가져온 설정이 병합됩니다.\\n계속하시겠습니까?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        
        if self.prefs.import_settings(filename):
            QMessageBox.information(self, "완료", "설정을 가져왔습니다.\\n변경 사항을 적용하려면 프로그램을 재시작하세요.")
            self._refresh_presets()
        else:
            QMessageBox.critical(self, "오류", "설정 가져오기에 실패했습니다.")

    def _import_excel(self):
        if not HAS_OPENPYXL:
            QMessageBox.critical(self, "오류", "openpyxl 라이브러리가 설치되지 않았습니다.\\npip install openpyxl")
            return
            
        fname, _ = QFileDialog.getOpenFileName(self, "엑셀 파일 열기", "", "Excel Files (*.xlsx)")
        if not fname: return
        
        try:
            wb = openpyxl.load_workbook(fname)
            ws = wb.active
            # Assume Row 2: Origin, Dest, DepDate, RetDate, Adults
            origin = ws['A2'].value
            dest = ws['B2'].value
            dep = str(ws['C2'].value)
            ret = str(ws['D2'].value) if ws['D2'].value else None
            adults = ws['E2'].value
            
            if origin and dest:
                params = {
                    "origin": origin, "dest": dest, 
                    "dep": dep, "ret": ret, "adults": int(adults) if adults else 1
                }
                self.prefs.save_profile("엑셀 가져옴", params)
                QMessageBox.information(self, "완료", "'엑셀 가져옴' 프로필로 저장되었습니다.\\n검색 패널에서 불러오세요.")
        except Exception as e:
            QMessageBox.critical(self, "오류", f"엑셀 읽기 실패: {e}")

    def _export_excel(self):
        if not HAS_OPENPYXL:
            QMessageBox.critical(self, "오류", "openpyxl 라이브러리가 설치되지 않았습니다.\\npip install openpyxl")
            return
            
        # Get results from MainWindow (parent)
        main_win = self.parent()
        if not main_win or not hasattr(main_win, 'all_results') or not main_win.all_results:
            QMessageBox.warning(self, "오류", "내보낼 검색 결과가 없습니다.")
            return

        fname, _ = QFileDialog.getSaveFileName(self, "엑셀로 저장", "flight_results.xlsx", "Excel Files (*.xlsx)")
        if not fname: return
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "검색결과"
            
            # Header
            headers = ["항공사", "가격", "출발", "도착", "경유", "복귀 출발", "복귀 도착", "복귀 경유", "출처"]
            ws.append(headers)
            
            for f in main_win.all_results:
                row = [
                    f.airline, f.price, 
                    f.departure_time, f.arrival_time, f.stops,
                    getattr(f, 'return_departure_time', '-'), 
                    getattr(f, 'return_arrival_time', '-'), 
                    getattr(f, 'return_stops', '-'),
                    f.source
                ]
                ws.append(row)
                
            wb.save(fname)
            QMessageBox.information(self, "완료", "엑셀 파일로 저장되었습니다.")
            
        except Exception as e:
            QMessageBox.critical(self, "오류", f"엑셀 저장 실패: {e}")
