"""
Custom UI Components for Flight Bot
"""
import sys
import csv
import logging
from datetime import datetime
from typing import Any
from PyQt6.QtWidgets import (
    QWidget, QHBoxLayout, QVBoxLayout, QGridLayout, QLabel, QPushButton, QCheckBox,
    QSpinBox, QComboBox, QDateEdit, QTabWidget, QFrame,
    QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView,
    QMenu, QMessageBox, QFileDialog, QApplication, QTextEdit,
    QRadioButton, QButtonGroup, QInputDialog
)
from PyQt6.QtCore import Qt, pyqtSignal, pyqtSlot, QDate, QSettings
from PyQt6.QtGui import QColor, QFont, QTextCharFormat

# Try importing openpyxl
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    openpyxl = None
    HAS_OPENPYXL = False

import config

logger = logging.getLogger(__name__)

class ResultTable(QTableWidget):
    favorite_requested = pyqtSignal(int)  # row index
    
    def __init__(self):
        super().__init__()
        self.results_data = []  # Store flight results for access
        
        self.setColumnCount(9)
        self.setHorizontalHeaderLabels([
            "항공사", "가격", "가는편 출발", "가는편 도착", "경유",
            "오는편 출발", "오는편 도착", "경유", "출처"
        ])
        
        # 열 너비 설정: 내용에 맞게 자동 조절 + 마지막 열 스트레치
        header = self.horizontalHeader()
        if header is not None:
            header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        initial_widths = [220, 180, 110, 110, 95, 110, 110, 95, 140]
        for idx, width in enumerate(initial_widths):
            if idx == 8:
                if header is not None:
                    header.setSectionResizeMode(idx, QHeaderView.ResizeMode.Stretch)
            else:
                self.setColumnWidth(idx, width)
        
        # 최소 너비 설정 (HiDPI 대응)
        if header is not None:
            header.setMinimumSectionSize(60)
        
        v_header = self.verticalHeader()
        if v_header is not None:
            v_header.setVisible(False)
            v_header.setDefaultSectionSize(48)
        self.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.setAlternatingRowColors(True)
        self.setSortingEnabled(True)

        # 렌더링 시 재사용할 스타일 객체
        self._font_placeholder = QFont("Pretendard", 12)
        self._font_price = QFont("Pretendard", 11, QFont.Weight.Bold)
        self._font_highlight = QFont("Pretendard", 11, QFont.Weight.Bold)
        self._color_placeholder = QColor("#64748b")
        self._color_price_cheap = QColor("#22c55e")
        self._color_price_good = QColor("#4cc9f0")
        self._color_price_mid = QColor("#f59e0b")
        self._color_price_high = QColor("#ef4444")
        self._color_stops_direct = QColor("#22c55e")
        self._color_stops_layover = QColor("#94a3b8")
        self._highlight_color = QColor(34, 197, 94, 40)
        
        # 테이블 스타일
        self.setStyleSheet("""
            QTableWidget {
                font-size: 13px;
            }
            QHeaderView::section {
                font-size: 13px;
                font-weight: 600;
                padding: 8px 4px;
            }
        """)
        
        # Enable context menu
        self.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.customContextMenuRequested.connect(self._show_context_menu)

    def update_data(self, results):
        # 대량 업데이트 최적화
        self.setUpdatesEnabled(False)
        self.setSortingEnabled(False)
        self.results_data = results
        
        # Handle empty results - show placeholder
        if not results:
            self.setRowCount(1)
            placeholder_item = QTableWidgetItem("🔍 검색 결과가 없습니다. 검색 조건을 확인해주세요.")
            placeholder_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            placeholder_item.setForeground(self._color_placeholder)
            placeholder_item.setFont(self._font_placeholder)
            self.setSpan(0, 0, 1, 9)
            self.setItem(0, 0, placeholder_item)
            self.setRowHeight(0, 80)
            self.setSortingEnabled(False)
            self.setUpdatesEnabled(True)
            return
        
        # Clear any previous span
        self.clearSpans()
        self.setRowCount(len(results))
        
        # Calculate price range for color coding
        min_price = min(r.price for r in results)
        max_price = max(r.price for r in results)
        price_range = max_price - min_price if max_price > min_price else 1
        
        for i, flight in enumerate(results):
            # Store flight object in first column's data
            airline_str = flight.airline
            if hasattr(flight, 'return_airline') and flight.return_airline and flight.airline != flight.return_airline:
                airline_str = f"{flight.airline} + {flight.return_airline}"
            
            airline_item = QTableWidgetItem(airline_str)
            airline_item.setData(Qt.ItemDataRole.UserRole + 1, i)
            # 툴팁에 상세 정보 표시
            if hasattr(flight, 'return_airline') and flight.return_airline:
                 airline_item.setToolTip(f"가는편: {flight.airline}\\n오는편: {flight.return_airline}")
            self.setItem(i, 0, airline_item)
            
            # Price (Color-coded: green=cheap, red=expensive)
            # 국내선: 가는편/오는편 가격 분리 표시
            # Add best price badge for minimum price
            if flight.price == min_price:
                if hasattr(flight, 'outbound_price') and flight.outbound_price > 0:
                    price_text = f"🏆 {flight.price:,}원 ({flight.outbound_price:,}+{flight.return_price:,})"
                else:
                    price_text = f"🏆 {flight.price:,}원"
            elif hasattr(flight, 'outbound_price') and flight.outbound_price > 0:
                price_text = f"{flight.price:,}원 ({flight.outbound_price:,}+{flight.return_price:,})"
            else:
                price_text = f"{flight.price:,}원"
            
            price_item = QTableWidgetItem(price_text)
            price_item.setData(Qt.ItemDataRole.UserRole, flight.price)
            price_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            tooltip_lines = [f"기본가: {flight.price:,}원"]
            if hasattr(flight, 'outbound_price') and flight.outbound_price > 0:
                tooltip_lines.append(
                    f"구성: {flight.outbound_price:,}원 + {flight.return_price:,}원"
                )
            if getattr(flight, 'benefit_price', 0) > 0:
                tooltip_lines.append(f"혜택가: {flight.benefit_price:,}원")
            if getattr(flight, 'benefit_label', ''):
                tooltip_lines.append(f"혜택 정보: {flight.benefit_label}")
            price_item.setToolTip("\n".join(tooltip_lines))
            
            # Color coding based on price position
            ratio = (flight.price - min_price) / price_range if price_range else 0
            if ratio < 0.2:
                price_color = self._color_price_cheap  # Green - cheapest
            elif ratio < 0.5:
                price_color = self._color_price_good  # Cyan - good
            elif ratio < 0.8:
                price_color = self._color_price_mid  # Orange - moderate
            else:
                price_color = self._color_price_high  # Red - expensive
            
            price_item.setForeground(price_color)
            price_item.setFont(self._font_price)
            self.setItem(i, 1, price_item)
            
            # Outbound
            self._set_time_item(i, 2, flight.departure_time)
            self._set_time_item(i, 3, flight.arrival_time)
            
            # Stops - highlight direct flights
            stops_item = QTableWidgetItem("✈️ 직항" if not flight.stops else f"{flight.stops}회 경유")
            if not flight.stops:
                stops_item.setForeground(self._color_stops_direct)
            else:
                stops_item.setForeground(self._color_stops_layover)
            self.setItem(i, 4, stops_item)
            
            # Inbound
            if hasattr(flight, 'is_round_trip') and flight.is_round_trip:
                self._set_time_item(i, 5, flight.return_departure_time)
                self._set_time_item(i, 6, flight.return_arrival_time)
                ret_stops = QTableWidgetItem("✈️ 직항" if not flight.return_stops else f"{flight.return_stops}회 경유")
                if not flight.return_stops:
                    ret_stops.setForeground(self._color_stops_direct)
                else:
                    ret_stops.setForeground(self._color_stops_layover)
                self.setItem(i, 7, ret_stops)
            else:
                self.setItem(i, 5, QTableWidgetItem("-"))
                self.setItem(i, 6, QTableWidgetItem("-"))
                self.setItem(i, 7, QTableWidgetItem("-"))
                
            # Source
            self.setItem(i, 8, QTableWidgetItem(flight.source))
            
            # 최저가 행 배경색 강조 (더 눈에 띄게)
            if flight.price == min_price:
                for col in range(self.columnCount()):
                    item = self.item(i, col)
                    if item:
                        item.setBackground(self._highlight_color)
                        item.setFont(self._font_highlight)
            
        self.setSortingEnabled(True)
        self.setUpdatesEnabled(True)  # 렌더링 다시 활성화


    def _set_time_item(self, row, col, text):
        item = QTableWidgetItem(text)
        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        self.setItem(row, col, item)
    
    def _show_context_menu(self, pos):
        row = self.rowAt(pos.y())
        if row < 0:
            return
        
        menu = QMenu(self)
        
        # Add to favorites
        action_fav = menu.addAction("⭐ 즐겨찾기 추가")
        if action_fav is not None:
            action_fav.triggered.connect(lambda: self.favorite_requested.emit(row))
        
        menu.addSeparator()
        
        # Copy info
        action_copy = menu.addAction("📋 정보 복사")
        if action_copy is not None:
            action_copy.triggered.connect(lambda: self._copy_row_info(row))
        
        menu.addSeparator()
        
        # Export options (전체 결과)
        action_excel = menu.addAction("📊 Excel로 내보내기")
        if action_excel is not None:
            action_excel.triggered.connect(self.export_to_excel)
        
        action_csv = menu.addAction("📥 CSV로 내보내기")
        if action_csv is not None:
            action_csv.triggered.connect(self.export_to_csv)
        
        menu.exec(self.mapToGlobal(pos))
    
    def _copy_row_info(self, row):
        flight = self.get_flight_at_row(row)
        if not flight:
            return
        info = f"{flight.airline} | {flight.price:,}원 | {flight.departure_time}→{flight.arrival_time}"
        clipboard = QApplication.clipboard()
        if clipboard is not None:
            clipboard.setText(info)
    
    def export_to_excel(self):
        """검색 결과를 Excel 파일로 내보내기"""
        if not self.results_data:
            QMessageBox.warning(self, "경고", "내보낼 데이터가 없습니다.")
            return
        
        if not HAS_OPENPYXL:
            QMessageBox.warning(self, "경고", "openpyxl이 설치되지 않았습니다.\\npip install openpyxl")
            return
        
        filename, _ = QFileDialog.getSaveFileName(
            self, "Excel로 저장", 
            f"flight_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if not filename:
            return
        
        try:
            if openpyxl is None:
                raise RuntimeError("openpyxl is unavailable")
            wb = openpyxl.Workbook()
            ws = wb.active
            if ws is None:
                raise RuntimeError("worksheet initialization failed")
            ws.title = "검색 결과"
            
            # 헤더
            headers = [
                "항공사",
                "오는편 항공사",
                "가격",
                "혜택가",
                "혜택 정보",
                "가는편 출발",
                "가는편 도착",
                "경유",
                "오는편 출발",
                "오는편 도착",
                "경유",
                "출처",
                "가는편 가격",
                "오는편 가격",
            ]
            ws.append(headers)
            
            # 데이터
            for flight in self.results_data:
                row = [
                    flight.airline,
                    getattr(flight, 'return_airline', ''),
                    flight.price,
                    getattr(flight, 'benefit_price', 0),
                    getattr(flight, 'benefit_label', ''),
                    flight.departure_time,
                    flight.arrival_time,
                    flight.stops,
                    getattr(flight, 'return_departure_time', ''),
                    getattr(flight, 'return_arrival_time', ''),
                    getattr(flight, 'return_stops', 0),
                    flight.source,
                    getattr(flight, 'outbound_price', 0),
                    getattr(flight, 'return_price', 0)
                ]
                ws.append(row)
            
            # 열 너비 자동 조절
            from openpyxl.utils import get_column_letter
            for col_idx, col in enumerate(ws.columns, start=1):
                max_length = max(len(str(cell.value or '')) for cell in col)
                ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2
            
            wb.save(filename)
            QMessageBox.information(self, "완료", f"Excel 파일이 저장되었습니다:\\n{filename}")
        except Exception as e:
            QMessageBox.critical(self, "오류", f"저장 실패: {e}")
    
    def export_to_csv(self):
        """검색 결과를 CSV 파일로 내보내기"""
        if not self.results_data:
            QMessageBox.warning(self, "경고", "내보낼 데이터가 없습니다.")
            return
        
        filename, _ = QFileDialog.getSaveFileName(
            self, "CSV로 저장",
            f"flight_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            "CSV Files (*.csv)"
        )
        
        if not filename:
            return
        
        try:
            with open(filename, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow([
                    "항공사",
                    "오는편 항공사",
                    "가격",
                    "혜택가",
                    "혜택 정보",
                    "가는편 출발",
                    "가는편 도착",
                    "경유",
                    "오는편 출발",
                    "오는편 도착",
                    "경유",
                    "출처",
                    "가는편 가격",
                    "오는편 가격",
                ])
                for flight in self.results_data:
                    writer.writerow([
                        flight.airline,
                        getattr(flight, 'return_airline', ''),
                        flight.price,
                        getattr(flight, 'benefit_price', 0),
                        getattr(flight, 'benefit_label', ''),
                        flight.departure_time,
                        flight.arrival_time,
                        flight.stops,
                        getattr(flight, 'return_departure_time', ''),
                        getattr(flight, 'return_arrival_time', ''),
                        getattr(flight, 'return_stops', 0),
                        flight.source,
                        getattr(flight, 'outbound_price', 0),
                        getattr(flight, 'return_price', 0)
                    ])
            QMessageBox.information(self, "완료", f"CSV 파일이 저장되었습니다:\\n{filename}")
        except Exception as e:
            QMessageBox.critical(self, "오류", f"저장 실패: {e}")
    
    def get_flight_at_row(self, row):
        """Get flight data for the given visual row"""
        if 0 <= row < len(self.results_data):
            # Account for sorting - get original index from item data
            item = self.item(row, 0)
            if item:
                orig_idx = item.data(Qt.ItemDataRole.UserRole + 1)
                if orig_idx is not None and 0 <= orig_idx < len(self.results_data):
                    return self.results_data[orig_idx]
        return None
