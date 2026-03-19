"""Dialogs for Flight Bot"""
import sys
import logging
from datetime import datetime, timedelta
from typing import Any, cast
from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QComboBox,
    QCalendarWidget, QGroupBox, QListWidget, QListWidgetItem, QFrame,
    QMessageBox, QDateEdit, QSpinBox, QCheckBox, QScrollArea, QGridLayout,
    QWidget, QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView,
    QTabWidget, QFileDialog, QInputDialog
)
from PyQt6.QtCore import Qt, pyqtSignal, QDate, QSettings
from PyQt6.QtGui import QColor, QFont, QTextCharFormat, QAction

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    openpyxl = None
    HAS_OPENPYXL = False

import config
from ui.styles import MODERN_THEME
from ui.components_primitives import NoWheelSpinBox, NoWheelComboBox, NoWheelDateEdit
from ui.dialogs_base import _validate_route_and_dates

logger = logging.getLogger(__name__)
class SettingsDialog(QDialog):
    def __init__(self, parent=None, prefs=None, db=None):
        super().__init__(parent)
        self.prefs: Any = prefs
        self.db: Any = db if db is not None else getattr(parent, "db", None)
        self.setWindowTitle("⚙️ 설정 (Settings)")
        self.setMinimumSize(600, 500)  # Increased size for better content display
        self.setStyleSheet(MODERN_THEME)
        self._init_ui()
        
    def _init_ui(self):
        layout = QVBoxLayout(self)
        
        tabs = QTabWidget()
        tabs.addTab(self._create_general_tab(), "일반")
        tabs.addTab(self._create_dest_tab(), "목적지 관리")
        tabs.addTab(self._create_data_tab(), "데이터 관리")
        
        layout.addWidget(tabs)
        
        btn_close = QPushButton("닫기")
        btn_close.clicked.connect(self.accept)
        layout.addWidget(btn_close)
        
    def _create_general_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Preferred Time
        grp_time = QGroupBox("기본 선호 시간대")
        gt_layout = QHBoxLayout(grp_time)
        
        self.spin_start = QSpinBox()
        self.spin_start.setRange(0, 23)
        self.spin_end = QSpinBox()
        self.spin_end.setRange(1, 24)
        
        pt = self.prefs.get_preferred_time()
        self.spin_start.setValue(pt.get("departure_start", 0))
        self.spin_end.setValue(pt.get("departure_end", 24))
        
        gt_layout.addWidget(QLabel("출발:"))
        gt_layout.addWidget(self.spin_start)
        gt_layout.addWidget(QLabel("~"))
        gt_layout.addWidget(self.spin_end)
        gt_layout.addWidget(QLabel("시"))
        
        # Max Results
        grp_limit = QGroupBox("검색 결과 제한")
        gl_layout = QHBoxLayout(grp_limit)
        
        self.spin_limit = QSpinBox()
        self.spin_limit.setRange(50, 2000)
        self.spin_limit.setSingleStep(50)
        self.spin_limit.setValue(self.prefs.get_max_results())
        self.spin_limit.setSuffix(" 개")
        self.spin_limit.setToolTip("한 번의 검색에서 표시할 최대 결과 수 (기본: 1000)")
        
        gl_layout.addWidget(QLabel("최대 표시 개수:"))
        gl_layout.addWidget(self.spin_limit)
        gl_layout.addStretch()
        
        # Save Button (Combined)
        btn_save_time = QPushButton("설정 저장")
        btn_save_time.setFixedWidth(100)  # Increased width for visibility
        btn_save_time.setFixedHeight(30)
        btn_save_time.clicked.connect(self._save_time_pref)
        
        gt_layout.addWidget(btn_save_time)
        
        layout.addWidget(grp_time)
        layout.addWidget(grp_limit)

        # Alert auto-check settings
        grp_alert = QGroupBox("🔔 자동 가격 알림 점검")
        ga_layout = QHBoxLayout(grp_alert)
        self.chk_alert_auto = QCheckBox("자동 점검 활성화")
        auto_cfg = self.prefs.get_alert_auto_check()
        self.chk_alert_auto.setChecked(auto_cfg.get("enabled", False))
        self.spin_alert_interval = QSpinBox()
        self.spin_alert_interval.setRange(5, 1440)
        self.spin_alert_interval.setValue(auto_cfg.get("interval_min", 30))
        self.spin_alert_interval.setSuffix(" 분")
        btn_save_alert = QPushButton("자동점검 저장")
        btn_save_alert.clicked.connect(self._save_alert_auto_check)
        ga_layout.addWidget(self.chk_alert_auto)
        ga_layout.addWidget(QLabel("주기:"))
        ga_layout.addWidget(self.spin_alert_interval)
        ga_layout.addWidget(btn_save_alert)
        ga_layout.addStretch()
        layout.addWidget(grp_alert)

        # Diagnostics
        grp_diag = QGroupBox("🩺 진단 (최근 24시간)")
        gd_layout = QVBoxLayout(grp_diag)
        self.lbl_diag = QLabel("진단 데이터를 불러오는 중...")
        self.lbl_diag.setWordWrap(True)
        self.lbl_diag.setStyleSheet("font-size: 12px; color: #cbd5e1;")
        btn_refresh_diag = QPushButton("진단 새로고침")
        btn_refresh_diag.clicked.connect(self._refresh_diagnostics)
        gd_layout.addWidget(self.lbl_diag)
        gd_layout.addWidget(btn_refresh_diag)
        layout.addWidget(grp_diag)

        self._refresh_diagnostics()
        layout.addStretch()
        return widget

    def _save_time_pref(self):
        self.prefs.set_preferred_time(self.spin_start.value(), self.spin_end.value())
        self.prefs.set_max_results(self.spin_limit.value())
        QMessageBox.information(self, "저장", "설정이 저장되었습니다.")

    def _save_alert_auto_check(self):
        self.prefs.set_alert_auto_check(
            self.chk_alert_auto.isChecked(),
            self.spin_alert_interval.value(),
        )
        QMessageBox.information(self, "저장", "자동 알림 점검 설정이 저장되었습니다.")

    def _refresh_diagnostics(self):
        if not self.db:
            self.lbl_diag.setText("진단 데이터베이스를 사용할 수 없습니다.")
            return
        try:
            summary = self.db.get_telemetry_summary(hours=24)
            selector = self.db.get_selector_health()
            error_top = summary.get("top_errors", [])
            top_text = ", ".join(f"{e['error_code']}({e['count']})" for e in error_top[:3]) if error_top else "없음"
            self.lbl_diag.setText(
                f"성공률: {summary.get('success_rate', 0.0):.1f}% | "
                f"수동모드 전환률: {summary.get('manual_mode_rate', 0.0):.1f}%\n"
                f"총 이벤트: {summary.get('total_events', 0)} | 주요 오류코드: {top_text}\n"
                f"Selector Health: {selector.get('overall_success_rate', 0.0):.1f}% "
                f"(표본 {selector.get('sample_count', 0)}건)"
            )
        except Exception as e:
            self.lbl_diag.setText(f"진단 조회 실패: {e}")

    def _create_dest_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        layout.addWidget(QLabel("사용자 정의 프리셋 목록:"))
        self.list_presets = QListWidget()
        self._refresh_presets()
        layout.addWidget(self.list_presets)
        
        btn_del = QPushButton("선택 삭제")
        btn_del.clicked.connect(self._delete_preset)
        layout.addWidget(btn_del)
        
        return widget
        
    def _refresh_presets(self):
        self.list_presets.clear()
        presets = self.prefs.get_all_presets()
        for code, name in presets.items():
            if code not in config.AIRPORTS: # Only show custom ones
                self.list_presets.addItem(f"{code} - {name}")

    def _delete_preset(self):
        item = self.list_presets.currentItem()
        if item:
            code = item.text().split(' - ')[0]
            self.prefs.remove_preset(code)
            self._refresh_presets()

    def _create_data_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # 프로필 가져오기/내보내기 그룹
        grp_profile = QGroupBox("📦 설정 백업/복원")
        gp_layout = QVBoxLayout(grp_profile)
        
        btn_export_settings = QPushButton("💾 모든 설정 내보내기")
        btn_export_settings.setToolTip("프리셋, 프로필, 선호 시간대, 테마 등 모든 설정을 JSON 파일로 저장")
        btn_export_settings.clicked.connect(self._export_all_settings)
        
        btn_import_settings = QPushButton("📂 설정 가져오기")
        btn_import_settings.setToolTip("JSON 파일에서 설정을 불러와 현재 설정에 병합")
        btn_import_settings.clicked.connect(self._import_all_settings)
        
        gp_layout.addWidget(btn_export_settings)
        gp_layout.addWidget(btn_import_settings)
        layout.addWidget(grp_profile)
        
        # 엑셀 그룹
        grp_excel = QGroupBox("📊 엑셀 (Excel)")
        gl = QVBoxLayout(grp_excel)
        
        btn_import = QPushButton("📂 검색 조건 가져오기 (Import)")
        btn_import.clicked.connect(self._import_excel)
        
        label_info = QLabel("엑셀 파일 양식: Origin, Dest, DepDate(YYYYMMDD), RetDate, Adults")
        label_info.setStyleSheet("font-size: 11px; color: #aaa;")
        
        btn_export = QPushButton("💾 검색 결과 내보내기 (Export)")
        btn_export.clicked.connect(self._export_excel)
        
        gl.addWidget(btn_import)
        gl.addWidget(btn_export)
        gl.addWidget(label_info)
        
        layout.addWidget(grp_excel)
        layout.addStretch()
        return widget

    def _export_all_settings(self):
        """모든 설정을 JSON 파일로 내보내기"""
        from datetime import datetime
        filename, _ = QFileDialog.getSaveFileName(
            self, "설정 내보내기",
            f"flight_bot_settings_{datetime.now().strftime('%Y%m%d')}.json",
            "JSON Files (*.json)"
        )
        if not filename:
            return
        
        if self.prefs.export_all_settings(filename):
            QMessageBox.information(self, "완료", f"설정이 저장되었습니다:\\n{filename}")
        else:
            QMessageBox.critical(self, "오류", "설정 내보내기에 실패했습니다.")

    def _import_all_settings(self):
        """JSON 파일에서 설정 가져오기"""
        filename, _ = QFileDialog.getOpenFileName(
            self, "설정 가져오기",
            "",
            "JSON Files (*.json)"
        )
        if not filename:
            return
        
        reply = QMessageBox.question(
            self, "설정 가져오기",
            "현재 설정에 가져온 설정이 병합됩니다.\\n계속하시겠습니까?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        
        if self.prefs.import_settings(filename):
            QMessageBox.information(self, "완료", "설정을 가져왔습니다.\\n변경 사항을 적용하려면 프로그램을 재시작하세요.")
            self._refresh_presets()
        else:
            QMessageBox.critical(self, "오류", "설정 가져오기에 실패했습니다.")

    def _import_excel(self):
        if not HAS_OPENPYXL:
            QMessageBox.critical(self, "오류", "openpyxl 라이브러리가 설치되지 않았습니다.\\npip install openpyxl")
            return
            
        fname, _ = QFileDialog.getOpenFileName(self, "엑셀 파일 열기", "", "Excel Files (*.xlsx)")
        if not fname: return
        
        try:
            if openpyxl is None:
                raise RuntimeError("openpyxl is unavailable")
            wb = openpyxl.load_workbook(fname)
            ws = wb.active
            if ws is None:
                raise RuntimeError("worksheet initialization failed")
            # Assume Row 2: Origin, Dest, DepDate, RetDate, Adults
            origin = ws['A2'].value
            dest = ws['B2'].value
            dep = str(ws['C2'].value)
            ret = str(ws['D2'].value) if ws['D2'].value else None
            adults = ws['E2'].value
            
            if origin and dest:
                params = {
                    "origin": origin,
                    "dest": dest,
                    "dep": dep,
                    "ret": ret,
                    "adults": int(adults) if adults else 1,
                    "cabin_class": "ECONOMY",
                    "is_domestic": config.infer_is_domestic_route(origin, dest),
                }
                self.prefs.save_profile("엑셀 가져옴", params)
                QMessageBox.information(self, "완료", "'엑셀 가져옴' 프로필로 저장되었습니다.\\n검색 패널에서 불러오세요.")
        except Exception as e:
            QMessageBox.critical(self, "오류", f"엑셀 읽기 실패: {e}")

    def _export_excel(self):
        if not HAS_OPENPYXL:
            QMessageBox.critical(self, "오류", "openpyxl 라이브러리가 설치되지 않았습니다.\\npip install openpyxl")
            return
            
        # Get results from MainWindow (parent)
        main_win = cast(Any, self.parent())
        if not main_win or not hasattr(main_win, 'all_results') or not main_win.all_results:
            QMessageBox.warning(self, "오류", "내보낼 검색 결과가 없습니다.")
            return

        fname, _ = QFileDialog.getSaveFileName(self, "엑셀로 저장", "flight_results.xlsx", "Excel Files (*.xlsx)")
        if not fname: return
        
        try:
            if openpyxl is None:
                raise RuntimeError("openpyxl is unavailable")
            wb = openpyxl.Workbook()
            ws = wb.active
            if ws is None:
                raise RuntimeError("worksheet initialization failed")
            ws.title = "검색결과"
            
            # Header
            headers = ["항공사", "가격", "출발", "도착", "경유", "복귀 출발", "복귀 도착", "복귀 경유", "출처"]
            ws.append(headers)
            
            for f in main_win.all_results:
                row = [
                    f.airline, f.price, 
                    f.departure_time, f.arrival_time, f.stops,
                    getattr(f, 'return_departure_time', '-'), 
                    getattr(f, 'return_arrival_time', '-'), 
                    getattr(f, 'return_stops', '-'),
                    f.source
                ]
                ws.append(row)
                
            wb.save(fname)
            QMessageBox.information(self, "완료", "엑셀 파일로 저장되었습니다.")
            
        except Exception as e:
            QMessageBox.critical(self, "오류", f"엑셀 저장 실패: {e}")
