"""Dialogs for Flight Bot"""
import sys
import logging
from datetime import datetime, timedelta
from typing import Any, cast
from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QComboBox,
    QCalendarWidget, QGroupBox, QListWidget, QListWidgetItem, QFrame,
    QMessageBox, QDateEdit, QSpinBox, QCheckBox, QScrollArea, QGridLayout,
    QWidget, QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView,
    QTabWidget, QFileDialog, QInputDialog
)
from PyQt6.QtCore import Qt, pyqtSignal, QDate, QSettings
from PyQt6.QtGui import QColor, QFont, QTextCharFormat, QAction

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    openpyxl = None
    HAS_OPENPYXL = False

import config
from ui.styles import MODERN_THEME
from ui.components_primitives import NoWheelSpinBox, NoWheelComboBox, NoWheelDateEdit
from ui.dialogs_base import _validate_route_and_dates

logger = logging.getLogger(__name__)

class ShortcutsDialog(QDialog):
    """키보드 단축키 도움말 다이얼로그"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("⌨️ 키보드 단축키")
        self.setMinimumSize(400, 300)
        self.setStyleSheet(MODERN_THEME)
        self._init_ui()
    
    def _init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        
        title = QLabel("키보드 단축키 목록")
        title.setStyleSheet("font-size: 18px; font-weight: bold; color: #4cc9f0;")
        layout.addWidget(title)
        
        shortcuts = [
            ("Ctrl + Enter", "검색 시작"),
            ("F5", "결과 새로고침 (필터 재적용)"),
            ("Escape", "검색 취소 / 다이얼로그 닫기"),
            ("Ctrl + F", "필터 항공사 선택으로 이동"),
            ("더블클릭", "결과 행에서 예약 페이지 열기"),
            ("우클릭", "결과 행에서 컨텍스트 메뉴"),
        ]
        
        for key, desc in shortcuts:
            row = QHBoxLayout()
            key_label = QLabel(key)
            key_label.setStyleSheet("""
                background-color: #0f3460;
                color: #4cc9f0;
                padding: 5px 10px;
                border-radius: 4px;
                font-family: 'Consolas', monospace;
                font-weight: bold;
            """)
            key_label.setFixedWidth(120)
            row.addWidget(key_label)
            
            desc_label = QLabel(desc)
            desc_label.setStyleSheet("color: #e6e6e6;")
            row.addWidget(desc_label)
            row.addStretch()
            
            layout.addLayout(row)
        
        layout.addStretch()
        
        btn_close = QPushButton("닫기")
        btn_close.clicked.connect(self.accept)
        layout.addWidget(btn_close)

class PriceAlertDialog(QDialog):
    """가격 알림 설정 다이얼로그"""
    
    def __init__(self, parent=None, db=None, prefs=None):
        super().__init__(parent)
        self.db: Any = db
        self.prefs: Any = prefs
        self.setWindowTitle("🔔 가격 알림 관리")
        self.setMinimumSize(700, 550)
        self.setStyleSheet(MODERN_THEME)
        self._init_ui()
        self._refresh_alerts()
    
    def _init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        
        # 설명
        info = QLabel("목표 가격을 설정하면 가격이 그 이하로 떨어질 때 알림을 받을 수 있습니다.")
        info.setStyleSheet("color: #94a3b8; font-size: 12px;")
        layout.addWidget(info)
        
        # 새 알림 추가 그룹
        grp_new = QGroupBox("➕ 새 알림 추가")
        new_layout = QGridLayout(grp_new)
        
        # 출발지
        new_layout.addWidget(QLabel("출발지:"), 0, 0)
        self.cb_origin = QComboBox()
        for code, name in config.AIRPORTS.items():
            self.cb_origin.addItem(f"{code} ({name})", code)
        new_layout.addWidget(self.cb_origin, 0, 1)
        
        # 도착지
        new_layout.addWidget(QLabel("도착지:"), 0, 2)
        self.cb_dest = QComboBox()
        all_presets = self.prefs.get_all_presets() if self.prefs else config.AIRPORTS
        for code, name in all_presets.items():
            self.cb_dest.addItem(f"{code} ({name})", code)
        self.cb_dest.setCurrentIndex(1)
        new_layout.addWidget(self.cb_dest, 0, 3)
        
        # 가는 날
        new_layout.addWidget(QLabel("가는 날:"), 1, 0)
        self.date_dep = QDateEdit()
        self.date_dep.setCalendarPopup(True)
        self.date_dep.setDate(QDate.currentDate().addDays(7))
        new_layout.addWidget(self.date_dep, 1, 1)
        
        # 오는 날
        new_layout.addWidget(QLabel("오는 날:"), 1, 2)
        self.date_ret = QDateEdit()
        self.date_ret.setCalendarPopup(True)
        self.date_ret.setDate(QDate.currentDate().addDays(10))
        new_layout.addWidget(self.date_ret, 1, 3)

        # 편도 알림
        self.chk_oneway = QCheckBox("편도 알림")
        self.chk_oneway.setToolTip("체크하면 귀국일 없이 편도 노선만 기준으로 알림을 설정합니다.")
        self.chk_oneway.toggled.connect(self._toggle_alert_oneway)
        new_layout.addWidget(self.chk_oneway, 2, 0, 1, 2)
        
        # 목표 가격
        new_layout.addWidget(QLabel("목표 가격:"), 3, 0)
        self.spin_target = QSpinBox()
        self.spin_target.setRange(10000, 10000000)
        self.spin_target.setSingleStep(10000)
        self.spin_target.setValue(300000)
        self.spin_target.setSuffix(" 원")
        new_layout.addWidget(self.spin_target, 3, 1)

        # 좌석 등급
        new_layout.addWidget(QLabel("좌석 등급:"), 2, 2)
        self.cb_cabin_class = QComboBox()
        self.cb_cabin_class.addItem("💺 이코노미", "ECONOMY")
        self.cb_cabin_class.addItem("💼 비즈니스", "BUSINESS")
        self.cb_cabin_class.addItem("👑 일등석", "FIRST")
        new_layout.addWidget(self.cb_cabin_class, 2, 3)
        
        # 추가 버튼
        btn_add = QPushButton("🔔 알림 추가")
        btn_add.clicked.connect(self._add_alert)
        new_layout.addWidget(btn_add, 3, 2, 1, 2)
        
        layout.addWidget(grp_new)
        
        # 현재 알림 목록
        alert_label = QLabel("📋 설정된 알림 목록:")
        alert_label.setObjectName("section_title")
        layout.addWidget(alert_label)
        
        self.table = QTableWidget()
        self.table.setColumnCount(8)
        self.table.setHorizontalHeaderLabels([
            "ID", "노선", "출발일", "귀국일", "좌석", "목표가", "현재가", "상태"
        ])
        self.table.setColumnHidden(0, True)  # ID 숨김
        header = self.table.horizontalHeader()
        if header is not None:
            header.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        layout.addWidget(self.table)
        
        # 버튼 영역
        btn_layout = QHBoxLayout()
        
        btn_refresh = QPushButton("🔄 새로고침")
        btn_refresh.clicked.connect(self._refresh_alerts)
        btn_layout.addWidget(btn_refresh)
        
        btn_delete = QPushButton("🗑️ 선택 삭제")
        btn_delete.clicked.connect(self._delete_selected)
        btn_layout.addWidget(btn_delete)
        
        btn_layout.addStretch()
        
        btn_close = QPushButton("닫기")
        btn_close.clicked.connect(self.accept)
        btn_layout.addWidget(btn_close)
        
        layout.addLayout(btn_layout)

    def _toggle_alert_oneway(self, checked: bool):
        self.date_ret.setEnabled(not checked)
    
    def _add_alert(self):
        """새 알림 추가"""
        origin = self.cb_origin.currentData()
        dest = self.cb_dest.currentData()
        dep_date = self.date_dep.date()
        ret_date = None if self.chk_oneway.isChecked() else self.date_ret.date()
        target = self.spin_target.value()
        cabin_class = self.cb_cabin_class.currentData() or "ECONOMY"

        if not _validate_route_and_dates(self, origin, dest, dep_date, ret_date):
            return

        dep = dep_date.toString("yyyyMMdd")
        ret = ret_date.toString("yyyyMMdd") if ret_date is not None else None

        try:
            alert_id = self.db.add_price_alert(origin, dest, dep, ret, target, cabin_class)
            QMessageBox.information(self, "완료", f"가격 알림이 추가되었습니다. (ID: {alert_id})")
            self._refresh_alerts()
        except Exception as e:
            QMessageBox.critical(self, "오류", f"알림 추가 실패: {e}")
    
    def _refresh_alerts(self):
        """알림 목록 새로고침"""
        if not self.db:
            return
        
        alerts = self.db.get_all_alerts()
        self.table.setRowCount(len(alerts))
        
        for i, alert in enumerate(alerts):
            self.table.setItem(i, 0, QTableWidgetItem(str(alert.id)))
            
            route = f"{alert.origin} → {alert.destination}"
            self.table.setItem(i, 1, QTableWidgetItem(route))
            
            # 날짜 포맷
            try:
                dep_dt = datetime.strptime(alert.departure_date, "%Y%m%d")
                dep_str = dep_dt.strftime("%Y-%m-%d")
            except:
                dep_str = alert.departure_date
            self.table.setItem(i, 2, QTableWidgetItem(dep_str))
            
            if alert.return_date:
                try:
                    ret_dt = datetime.strptime(alert.return_date, "%Y%m%d")
                    ret_str = ret_dt.strftime("%Y-%m-%d")
                except:
                    ret_str = alert.return_date
            else:
                ret_str = "-"
            self.table.setItem(i, 3, QTableWidgetItem(ret_str))

            cabin_class = getattr(alert, "cabin_class", "ECONOMY") or "ECONOMY"
            cabin_text = {"ECONOMY": "이코노미", "BUSINESS": "비즈니스", "FIRST": "일등석"}.get(cabin_class, cabin_class)
            self.table.setItem(i, 4, QTableWidgetItem(cabin_text))
            
            # 목표 가격
            target_item = QTableWidgetItem(f"{alert.target_price:,}원")
            target_item.setForeground(QColor("#4cc9f0"))
            self.table.setItem(i, 5, target_item)
            
            # 현재 가격
            if alert.last_price:
                current_item = QTableWidgetItem(f"{alert.last_price:,}원")
                if alert.last_price <= alert.target_price:
                    current_item.setForeground(QColor("#22c55e"))
                else:
                    current_item.setForeground(QColor("#f59e0b"))
            else:
                current_item = QTableWidgetItem("미확인")
            self.table.setItem(i, 6, current_item)
            
            # 상태
            if alert.triggered:
                status = "✅ 발동됨"
                color = "#22c55e"
            elif alert.is_active:
                status = "🔔 활성"
                color = "#4cc9f0"
            else:
                status = "⏸️ 비활성"
                color = "#94a3b8"
            
            status_item = QTableWidgetItem(status)
            status_item.setForeground(QColor(color))
            self.table.setItem(i, 7, status_item)
    
    def _delete_selected(self):
        """선택된 알림 삭제"""
        row = self.table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "선택 오류", "삭제할 알림을 선택하세요.")
            return
        
        alert_id_item = self.table.item(row, 0)
        if alert_id_item is None:
            QMessageBox.warning(self, "선택 오류", "선택된 알림 정보를 읽을 수 없습니다.")
            return
        alert_id = int(alert_id_item.text())
        reply = QMessageBox.question(
            self, "삭제 확인",
            "선택한 알림을 삭제하시겠습니까?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            if self.db.delete_alert(alert_id):
                self._refresh_alerts()
            else:
                QMessageBox.critical(self, "오류", "알림 삭제에 실패했습니다.")

class SettingsDialog(QDialog):
    def __init__(self, parent=None, prefs=None, db=None):
        super().__init__(parent)
        self.prefs: Any = prefs
        self.db: Any = db if db is not None else getattr(parent, "db", None)
        self.setWindowTitle("⚙️ 설정 (Settings)")
        self.setMinimumSize(600, 500)  # Increased size for better content display
        self.setStyleSheet(MODERN_THEME)
        self._init_ui()
        
    def _init_ui(self):
        layout = QVBoxLayout(self)
        
        tabs = QTabWidget()
        tabs.addTab(self._create_general_tab(), "일반")
        tabs.addTab(self._create_dest_tab(), "목적지 관리")
        tabs.addTab(self._create_data_tab(), "데이터 관리")
        
        layout.addWidget(tabs)
        
        btn_close = QPushButton("닫기")
        btn_close.clicked.connect(self.accept)
        layout.addWidget(btn_close)
        
    def _create_general_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Preferred Time
        grp_time = QGroupBox("기본 선호 시간대")
        gt_layout = QHBoxLayout(grp_time)
        
        self.spin_start = QSpinBox()
        self.spin_start.setRange(0, 23)
        self.spin_end = QSpinBox()
        self.spin_end.setRange(1, 24)
        
        pt = self.prefs.get_preferred_time()
        self.spin_start.setValue(pt.get("departure_start", 0))
        self.spin_end.setValue(pt.get("departure_end", 24))
        
        gt_layout.addWidget(QLabel("출발:"))
        gt_layout.addWidget(self.spin_start)
        gt_layout.addWidget(QLabel("~"))
        gt_layout.addWidget(self.spin_end)
        gt_layout.addWidget(QLabel("시"))
        
        # Max Results
        grp_limit = QGroupBox("검색 결과 제한")
        gl_layout = QHBoxLayout(grp_limit)
        
        self.spin_limit = QSpinBox()
        self.spin_limit.setRange(50, 2000)
        self.spin_limit.setSingleStep(50)
        self.spin_limit.setValue(self.prefs.get_max_results())
        self.spin_limit.setSuffix(" 개")
        self.spin_limit.setToolTip("한 번의 검색에서 표시할 최대 결과 수 (기본: 1000)")
        
        gl_layout.addWidget(QLabel("최대 표시 개수:"))
        gl_layout.addWidget(self.spin_limit)
        gl_layout.addStretch()
        
        # Save Button (Combined)
        btn_save_time = QPushButton("설정 저장")
        btn_save_time.setFixedWidth(100)  # Increased width for visibility
        btn_save_time.setFixedHeight(30)
        btn_save_time.clicked.connect(self._save_time_pref)
        
        gt_layout.addWidget(btn_save_time)
        
        layout.addWidget(grp_time)
        layout.addWidget(grp_limit)

        # Alert auto-check settings
        grp_alert = QGroupBox("🔔 자동 가격 알림 점검")
        ga_layout = QHBoxLayout(grp_alert)
        self.chk_alert_auto = QCheckBox("자동 점검 활성화")
        auto_cfg = self.prefs.get_alert_auto_check()
        self.chk_alert_auto.setChecked(auto_cfg.get("enabled", False))
        self.spin_alert_interval = QSpinBox()
        self.spin_alert_interval.setRange(5, 1440)
        self.spin_alert_interval.setValue(auto_cfg.get("interval_min", 30))
        self.spin_alert_interval.setSuffix(" 분")
        btn_save_alert = QPushButton("자동점검 저장")
        btn_save_alert.clicked.connect(self._save_alert_auto_check)
        ga_layout.addWidget(self.chk_alert_auto)
        ga_layout.addWidget(QLabel("주기:"))
        ga_layout.addWidget(self.spin_alert_interval)
        ga_layout.addWidget(btn_save_alert)
        ga_layout.addStretch()
        layout.addWidget(grp_alert)

        # Diagnostics
        grp_diag = QGroupBox("🩺 진단 (최근 24시간)")
        gd_layout = QVBoxLayout(grp_diag)
        self.lbl_diag = QLabel("진단 데이터를 불러오는 중...")
        self.lbl_diag.setWordWrap(True)
        self.lbl_diag.setStyleSheet("font-size: 12px; color: #cbd5e1;")
        btn_refresh_diag = QPushButton("진단 새로고침")
        btn_refresh_diag.clicked.connect(self._refresh_diagnostics)
        gd_layout.addWidget(self.lbl_diag)
        gd_layout.addWidget(btn_refresh_diag)
        layout.addWidget(grp_diag)

        self._refresh_diagnostics()
        layout.addStretch()
        return widget

    def _save_time_pref(self):
        self.prefs.set_preferred_time(self.spin_start.value(), self.spin_end.value())
        self.prefs.set_max_results(self.spin_limit.value())
        QMessageBox.information(self, "저장", "설정이 저장되었습니다.")

    def _save_alert_auto_check(self):
        self.prefs.set_alert_auto_check(
            self.chk_alert_auto.isChecked(),
            self.spin_alert_interval.value(),
        )
        QMessageBox.information(self, "저장", "자동 알림 점검 설정이 저장되었습니다.")

    def _refresh_diagnostics(self):
        if not self.db:
            self.lbl_diag.setText("진단 데이터베이스를 사용할 수 없습니다.")
            return
        try:
            summary = self.db.get_telemetry_summary(hours=24)
            selector = self.db.get_selector_health()
            error_top = summary.get("top_errors", [])
            top_text = ", ".join(f"{e['error_code']}({e['count']})" for e in error_top[:3]) if error_top else "없음"
            self.lbl_diag.setText(
                f"성공률: {summary.get('success_rate', 0.0):.1f}% | "
                f"수동모드 전환률: {summary.get('manual_mode_rate', 0.0):.1f}%\n"
                f"총 이벤트: {summary.get('total_events', 0)} | 주요 오류코드: {top_text}\n"
                f"Selector Health: {selector.get('overall_success_rate', 0.0):.1f}% "
                f"(표본 {selector.get('sample_count', 0)}건)"
            )
        except Exception as e:
            self.lbl_diag.setText(f"진단 조회 실패: {e}")

    def _create_dest_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        layout.addWidget(QLabel("사용자 정의 프리셋 목록:"))
        self.list_presets = QListWidget()
        self._refresh_presets()
        layout.addWidget(self.list_presets)
        
        btn_del = QPushButton("선택 삭제")
        btn_del.clicked.connect(self._delete_preset)
        layout.addWidget(btn_del)
        
        return widget
        
    def _refresh_presets(self):
        self.list_presets.clear()
        presets = self.prefs.get_all_presets()
        for code, name in presets.items():
            if code not in config.AIRPORTS: # Only show custom ones
                self.list_presets.addItem(f"{code} - {name}")

    def _delete_preset(self):
        item = self.list_presets.currentItem()
        if item:
            code = item.text().split(' - ')[0]
            self.prefs.remove_preset(code)
            self._refresh_presets()

    def _create_data_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # 프로필 가져오기/내보내기 그룹
        grp_profile = QGroupBox("📦 설정 백업/복원")
        gp_layout = QVBoxLayout(grp_profile)
        
        btn_export_settings = QPushButton("💾 모든 설정 내보내기")
        btn_export_settings.setToolTip("프리셋, 프로필, 선호 시간대, 테마 등 모든 설정을 JSON 파일로 저장")
        btn_export_settings.clicked.connect(self._export_all_settings)
        
        btn_import_settings = QPushButton("📂 설정 가져오기")
        btn_import_settings.setToolTip("JSON 파일에서 설정을 불러와 현재 설정에 병합")
        btn_import_settings.clicked.connect(self._import_all_settings)
        
        gp_layout.addWidget(btn_export_settings)
        gp_layout.addWidget(btn_import_settings)
        layout.addWidget(grp_profile)
        
        # 엑셀 그룹
        grp_excel = QGroupBox("📊 엑셀 (Excel)")
        gl = QVBoxLayout(grp_excel)
        
        btn_import = QPushButton("📂 검색 조건 가져오기 (Import)")
        btn_import.clicked.connect(self._import_excel)
        
        label_info = QLabel("엑셀 파일 양식: Origin, Dest, DepDate(YYYYMMDD), RetDate, Adults")
        label_info.setStyleSheet("font-size: 11px; color: #aaa;")
        
        btn_export = QPushButton("💾 검색 결과 내보내기 (Export)")
        btn_export.clicked.connect(self._export_excel)
        
        gl.addWidget(btn_import)
        gl.addWidget(btn_export)
        gl.addWidget(label_info)
        
        layout.addWidget(grp_excel)
        layout.addStretch()
        return widget

    def _export_all_settings(self):
        """모든 설정을 JSON 파일로 내보내기"""
        from datetime import datetime
        filename, _ = QFileDialog.getSaveFileName(
            self, "설정 내보내기",
            f"flight_bot_settings_{datetime.now().strftime('%Y%m%d')}.json",
            "JSON Files (*.json)"
        )
        if not filename:
            return
        
        if self.prefs.export_all_settings(filename):
            QMessageBox.information(self, "완료", f"설정이 저장되었습니다:\\n{filename}")
        else:
            QMessageBox.critical(self, "오류", "설정 내보내기에 실패했습니다.")

    def _import_all_settings(self):
        """JSON 파일에서 설정 가져오기"""
        filename, _ = QFileDialog.getOpenFileName(
            self, "설정 가져오기",
            "",
            "JSON Files (*.json)"
        )
        if not filename:
            return
        
        reply = QMessageBox.question(
            self, "설정 가져오기",
            "현재 설정에 가져온 설정이 병합됩니다.\\n계속하시겠습니까?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        
        if self.prefs.import_settings(filename):
            QMessageBox.information(self, "완료", "설정을 가져왔습니다.\\n변경 사항을 적용하려면 프로그램을 재시작하세요.")
            self._refresh_presets()
        else:
            QMessageBox.critical(self, "오류", "설정 가져오기에 실패했습니다.")

    def _import_excel(self):
        if not HAS_OPENPYXL:
            QMessageBox.critical(self, "오류", "openpyxl 라이브러리가 설치되지 않았습니다.\\npip install openpyxl")
            return
            
        fname, _ = QFileDialog.getOpenFileName(self, "엑셀 파일 열기", "", "Excel Files (*.xlsx)")
        if not fname: return
        
        try:
            if openpyxl is None:
                raise RuntimeError("openpyxl is unavailable")
            wb = openpyxl.load_workbook(fname)
            ws = wb.active
            if ws is None:
                raise RuntimeError("worksheet initialization failed")
            # Assume Row 2: Origin, Dest, DepDate, RetDate, Adults
            origin = ws['A2'].value
            dest = ws['B2'].value
            dep = str(ws['C2'].value)
            ret = str(ws['D2'].value) if ws['D2'].value else None
            adults = ws['E2'].value
            
            if origin and dest:
                params = {
                    "origin": origin, "dest": dest, 
                    "dep": dep, "ret": ret, "adults": int(adults) if adults else 1
                }
                self.prefs.save_profile("엑셀 가져옴", params)
                QMessageBox.information(self, "완료", "'엑셀 가져옴' 프로필로 저장되었습니다.\\n검색 패널에서 불러오세요.")
        except Exception as e:
            QMessageBox.critical(self, "오류", f"엑셀 읽기 실패: {e}")

    def _export_excel(self):
        if not HAS_OPENPYXL:
            QMessageBox.critical(self, "오류", "openpyxl 라이브러리가 설치되지 않았습니다.\\npip install openpyxl")
            return
            
        # Get results from MainWindow (parent)
        main_win = cast(Any, self.parent())
        if not main_win or not hasattr(main_win, 'all_results') or not main_win.all_results:
            QMessageBox.warning(self, "오류", "내보낼 검색 결과가 없습니다.")
            return

        fname, _ = QFileDialog.getSaveFileName(self, "엑셀로 저장", "flight_results.xlsx", "Excel Files (*.xlsx)")
        if not fname: return
        
        try:
            if openpyxl is None:
                raise RuntimeError("openpyxl is unavailable")
            wb = openpyxl.Workbook()
            ws = wb.active
            if ws is None:
                raise RuntimeError("worksheet initialization failed")
            ws.title = "검색결과"
            
            # Header
            headers = ["항공사", "가격", "출발", "도착", "경유", "복귀 출발", "복귀 도착", "복귀 경유", "출처"]
            ws.append(headers)
            
            for f in main_win.all_results:
                row = [
                    f.airline, f.price, 
                    f.departure_time, f.arrival_time, f.stops,
                    getattr(f, 'return_departure_time', '-'), 
                    getattr(f, 'return_arrival_time', '-'), 
                    getattr(f, 'return_stops', '-'),
                    f.source
                ]
                ws.append(row)
                
            wb.save(fname)
            QMessageBox.information(self, "완료", "엑셀 파일로 저장되었습니다.")
            
        except Exception as e:
            QMessageBox.critical(self, "오류", f"엑셀 저장 실패: {e}")
